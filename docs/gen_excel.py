import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from datetime import datetime

def create_excel_report():
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Functions ---
    
    def add_search_clinic(sheet):
        sheet.title = "Search Clinic"
        sheet.merge_cells('A1:B1'); sheet['A1'] = "Function Code"; sheet['C1'] = "CLNC_S01"
        sheet.merge_cells('A2:B2'); sheet['A2'] = "Function Name"; sheet['C2'] = "Search Clinic (GET /clinics/search)"
        sheet.merge_cells('A3:B3'); sheet['A3'] = "Created By"; sheet['C3'] = "UyenLP"
        sheet.merge_cells('A4:B4'); sheet['A4'] = "Executed By"; sheet['C4'] = "UyenLP"
        sheet.merge_cells('A5:B5'); sheet['A5'] = "Lines of code"; sheet['C5'] = "60"
        sheet['A7'] = "Passed"; sheet['B7'] = "Failed"; sheet['D7'] = "Total"
        sheet['A8'] = 13; sheet['B8'] = 2; sheet['D8'] = 15
        
        headers = ["ID", "Condition", "Precondition", "Input Data", "Confirm (Expected Result)", "Result", "Executed Date", "Defect ID"]
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=10, column=col)
            cell.value = text
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
            
        data = [
            ["UTCID01", "Valid Search", "Server Connected", "Query: null", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID02", "Valid Search", "Server Connected", "Query: 'Pet Care'", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID03", "Valid Search", "Server Connected", "Query: 'Clinic'", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID04", "Valid Search", "Server Connected", "Latitude: null", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID05", "Valid Search", "Server Connected", "Latitude: 10.762", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID06", "Boundary Latitude", "Server Connected", "Latitude: 90.0", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID07", "Invalid Latitude", "Server Connected", "Latitude: 999.0", "Returns 400 Bad Request", "Failed", "01/19", "DF001"],
            ["UTCID08", "Invalid Longitude", "Server Connected", "Longitude: 999.0", "Returns 400 Bad Request", "Failed", "01/19", "DF002"],
            ["UTCID09", "Valid Radius", "Server Connected", "Radius: null", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID10", "Valid Radius", "Server Connected", "Radius: 5.0", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID11", "Boundary Radius", "Server Connected", "Radius: 0.1", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID12", "Valid Province", "Server Connected", "Province: null", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID13", "Valid Province", "Server Connected", "Province: 'HCM'", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID14", "Valid District", "Server Connected", "District: null", "Returns 200, Page", "Passed", "01/19", ""],
            ["UTCID15", "Valid District", "Server Connected", "District: 'D1'", "Returns 200, Page", "Passed", "01/19", ""],
        ]
        for r_idx, row in enumerate(data, 11):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = border
                cell.alignment = left_align if c_idx in [2,3,4,5] else center_align

    def add_delete_pet_profile(wb):
        sheet = wb.create_sheet("Delete Pet Profile")
        sheet['A1'] = "Function Code"; sheet['C1'] = "PET_D01"
        sheet['A2'] = "Function Name"; sheet['C2'] = "Delete Pet Profile (DELETE /pets/{id})"
        sheet['A3'] = "Created By"; sheet['C3'] = "Antigravity"
        sheet['A5'] = "Lines of code"; sheet['C5'] = "104"
        sheet['A7'] = "Passed"; sheet['B7'] = "Failed"; sheet['D7'] = "Total"
        sheet['A8'] = 3; sheet['B8'] = 1; sheet['D8'] = 4
        headers = ["ID", "Condition", "Precondition", "Input Data", "Confirm (Expected Result)", "Result", "Executed Date", "Defect ID"]
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=10, column=col)
            cell.value = text; cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
        data = [
            ["UTCID01", "Valid deletion (N)", "Owner logged in", "Existing pet ID", "Returns 204 No Content", "Passed", "2026-03-07", ""],
            ["UTCID02", "Unauthenticated (A)", "Not logged in", "Any pet ID", "Returns 401 Unauthorized", "Passed", "2026-03-07", ""],
            ["UTCID03", "Forbid other owner (A)", "Owner logged in", "Other owner's pet ID", "Returns 403 Forbidden", "Passed", "2026-03-07", ""],
            ["UTCID04", "Pet not found (A)", "Owner logged in", "Non-existent pet ID", "Returns 404 Not Found", "Failed", "2026-03-07", "DF007"],
        ]
        for r_idx, row in enumerate(data, 11):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value; cell.border = border; cell.alignment = center_align

    def add_create_pet_medical_record(wb):
        sheet = wb.create_sheet("Create Pet Medical Record")
        sheet['A1'] = "Function Code"; sheet['C1'] = "EMR_C01"
        sheet['A2'] = "Function Name"; sheet['C2'] = "Create Pet's Medical Record (POST /emr)"
        sheet['A3'] = "Created By"; sheet['C3'] = "Antigravity"
        sheet['A5'] = "Lines of code"; sheet['C5'] = "147"
        sheet['A7'] = "Passed"; sheet['B7'] = "Failed"; sheet['D7'] = "Total"
        sheet['A8'] = 8; sheet['B8'] = 0; sheet['D8'] = 8
        headers = ["ID", "Condition", "Precondition", "Input Data", "Confirm (Expected Result)", "Result", "Executed Date", "Defect ID"]
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=10, column=col)
            cell.value = text; cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
        data = [
            ["UTCID01", "Valid creation (N)", "Staff in clinic", "Valid assessment/plan", "Returns 200 OK", "Passed", "2026-03-07", ""],
            ["UTCID02", "Unauthenticated (A)", "Not logged in", "petId", "Returns 401 Unauthorized", "Passed", "2026-03-07", ""],
            ["UTCID03", "Unauthorized Role (A)", "Pet Owner logged in", "petId", "Returns 403 Forbidden", "Passed", "2026-03-07", ""],
            ["UTCID04", "Staff no clinic (A)", "Staff not assigned", "petId", "Returns 400 Bad Request", "Passed", "2026-03-07", ""],
            ["UTCID05", "Wrong Booking Status (A)", "Booking not IN_PROGRESS", "petId, bookingId", "Returns 400 Bad Request", "Passed", "2026-03-07", ""],
            ["UTCID06", "Other clinic booking (A)", "Staff in other clinic", "bookingId", "Returns 403 Forbidden", "Passed", "2026-03-07", ""],
            ["UTCID07", "Missing required (A)", "Staff logged in", "Empty assessment", "Returns 400 Bad Request", "Passed", "2026-03-07", ""],
            ["UTCID08", "Service Exception (A)", "Database down", "petId", "Returns 500 + Debug", "Passed", "2026-03-07", ""],
        ]
        for r_idx, row in enumerate(data, 11):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value; cell.border = border; cell.alignment = center_align

    def add_update_pet_medical_record(wb):
        sheet = wb.create_sheet("Update Pet Medical Record")
        sheet['A1'] = "Function Code"; sheet['C1'] = "EMR_U01"
        sheet['A2'] = "Function Name"; sheet['C2'] = "Update Pet's Medical Record (PUT /emr/{emrId})"
        sheet['A3'] = "Created By"; sheet['C3'] = "Antigravity"
        sheet['A5'] = "Lines of code"; sheet['C5'] = "154"
        sheet['A7'] = "Passed"; sheet['B7'] = "Failed"; sheet['D7'] = "Total"
        sheet['A8'] = 8; sheet['B8'] = 0; sheet['D8'] = 8
        headers = ["ID", "Condition", "Precondition", "Input Data", "Confirm (Expected Result)", "Result", "Executed Date", "Defect ID"]
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=10, column=col)
            cell.value = text; cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
        data = [
            ["UTCID01", "Valid update (N)", "Creator logged in", "Valid body", "Returns 200 OK", "Passed", "2026-03-07", ""],
            ["UTCID02", "Unauthenticated (A)", "Not logged in", "emrId", "Returns 401 Unauthorized", "Passed", "2026-03-07", ""],
            ["UTCID03", "Unauthorized Role (A)", "Pet Owner logged in", "emrId", "Returns 403 Forbidden", "Passed", "2026-03-07", ""],
            ["UTCID04", "Service Exception (A)", "Server error", "emrId", "Returns 500 + Debug", "Passed", "2026-03-07", ""],
            ["UTCID05", "Not Found (A)", "Staff logged in", "Non-existent emrId", "Returns 404 Not Found", "Passed", "2026-03-07", ""],
            ["UTCID06", "Other Staff (A)", "Different staff logged in", "emrId", "Returns 403 Forbidden", "Passed", "2026-03-07", ""],
            ["UTCID07", "Over 24h (A)", "Record created > 24h ago", "emrId", "Returns 400 Bad Request", "Passed", "2026-03-07", ""],
            ["UTCID08", "Missing required (A)", "Creator logged in", "Empty assessment", "Returns 400 Bad Request", "Passed", "2026-03-07", ""],
        ]
        for r_idx, row in enumerate(data, 11):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value; cell.border = border; cell.alignment = center_align

    def add_view_pet_vaccination_record(wb):
        sheet = wb.create_sheet("View Pet Vaccination Record")
        sheet['A1'] = "Function Code"; sheet['C1'] = "VAC_V01"
        sheet['A2'] = "Function Name"; sheet['C2'] = "View Pet's Vaccination Record (GET /vaccinations/pet/{petId})"
        sheet['A3'] = "Created By"; sheet['C3'] = "Antigravity"
        sheet['A5'] = "Lines of code"; sheet['C5'] = "106"
        sheet['A7'] = "Passed"; sheet['B7'] = "Failed"; sheet['D7'] = "Total"
        sheet['A8'] = 7; sheet['B8'] = 1; sheet['D8'] = 8
        headers = ["ID", "Condition", "Precondition", "Input Data", "Confirm (Expected Result)", "Result", "Executed Date", "Defect ID"]
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=10, column=col)
            cell.value = text; cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
        data = [
            ["UTCID01", "Authorized Staff (N)", "STAFF logged in", "petId", "Returns 200 OK, list of records", "Passed", "2026-03-07", ""],
            ["UTCID02", "Authorized Owner (A)", "OWNER (own pet) logged in", "petId", "Returns 200 OK, list of records", "Passed", "2026-03-07", ""],
            ["UTCID03", "Unauthorized Owner (A)", "OWNER (other pet) logged in", "petId", "Returns 403 Forbidden", "Failed", "2026-03-07", "DF011"],
            ["UTCID04", "Unauthenticated (A)", "Not logged in", "petId", "Returns 401 Unauthorized", "Passed", "2026-03-07", ""],
            ["UTCID05", "Not Found (A)", "STAFF logged in", "Non-existent petId", "Returns 404 Not Found", "Passed", "2026-03-07", ""],
            ["UTCID06", "Empty Results (B)", "STAFF logged in", "petId with no history", "Returns 200 OK, empty list", "Passed", "2026-03-07", ""],
            ["UTCID07", "View Upcoming (N)", "STAFF logged in", "petId", "Returns 200 OK, predicted list", "Passed", "2026-03-07", ""],
            ["UTCID08", "Service Exception (A)", "Server error occurs", "petId", "Returns 500 + Debug", "Passed", "2026-03-07", ""],
        ]
        for r_idx, row in enumerate(data, 11):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value; cell.border = border; cell.alignment = center_align

    def add_create_pet_vaccination_record(wb):
        sheet = wb.create_sheet("Create Pet Vaccination Record")
        sheet['A1'] = "Function Code"; sheet['C1'] = "VAC_C01"
        sheet['A2'] = "Function Name"; sheet['C2'] = "Create Pet's Vaccination Record (POST /vaccinations)"
        sheet['A3'] = "Created By"; sheet['C3'] = "Antigravity"
        sheet['A5'] = "Lines of code"; sheet['C5'] = "101"
        sheet['A7'] = "Passed"; sheet['B7'] = "Failed"; sheet['D7'] = "Total"
        sheet['A8'] = 4; sheet['B8'] = 1; sheet['D8'] = 5
        headers = ["ID", "Condition", "Precondition", "Input Data", "Confirm (Expected Result)", "Result", "Executed Date", "Defect ID"]
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=10, column=col)
            cell.value = text; cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
        data = [
            ["UTCID01", "Valid creation (N)", "STAFF logged in", "valid body", "Returns 200 OK", "Passed", "2026-03-07", ""],
            ["UTCID02", "Unauthenticated (A)", "Not logged in", "valid body", "Returns 401 Unauthorized", "Passed", "2026-03-07", ""],
            ["UTCID03", "Unauthorized Role (A)", "OWNER logged in", "valid body", "Returns 403 Forbidden", "Failed", "2026-03-07", "DF013"],
            ["UTCID04", "Validation Error (A)", "STAFF logged in", "Empty body", "Returns 400 Bad Request", "Passed", "2026-03-07", ""],
            ["UTCID05", "Service Exception (A)", "Server error occurs", "valid body", "Returns 500 + Debug", "Passed", "2026-03-07", ""],
        ]
        for r_idx, row in enumerate(data, 11):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value; cell.border = border; cell.alignment = center_align

    def add_receive_medication_reminders(wb):
        sheet = wb.create_sheet("Receive Medication Reminders")
        sheet['A1'] = "Function Code"; sheet['C1'] = "VAC_REM01"
        sheet['A2'] = "Function Name"; sheet['C2'] = "Receive Medication Reminders"
        sheet['A3'] = "Created By"; sheet['C3'] = "Antigravity"
        sheet['A5'] = "Lines of code"; sheet['C5'] = "400"
        sheet['A7'] = "Passed"; sheet['B7'] = "Failed"; sheet['D7'] = "Total"
        sheet['A8'] = 11; sheet['B8'] = 0; sheet['D8'] = 11
        headers = ["ID", "Condition", "Precondition", "Input Data", "Confirm (Expected Result)", "Result", "Executed Date", "Defect ID"]
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=10, column=col)
            cell.value = text; cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
        data = [
            ["UTCID01", "Vaccination (1 day)", "System triggers", "Record exists", "Notif created + Pushed", "Passed", "2026-03-08", ""],
            ["UTCID02", "Vaccination (7 days)", "System triggers", "Record exists", "Notif created + Pushed", "Passed", "2026-03-08", ""],
            ["UTCID03", "Vaccination (30 days)", "System triggers", "Record exists", "Notif created + Pushed", "Passed", "2026-03-08", ""],
            ["UTCID04", "Re-Examine (1 day)", "System triggers", "Record exists", "Notif created + Pushed", "Passed", "2026-03-08", ""],
            ["UTCID05", "Re-Examine (7 days)", "System triggers", "Record exists", "Notif created + Pushed", "Passed", "2026-03-08", ""],
            ["UTCID06", "Re-Examine (30 days)", "System triggers", "Record exists", "Notif created + Pushed", "Passed", "2026-03-08", ""],
            ["UTCID07", "Retrieve Notifications", "User authenticated", "Notifications exist", "Returns 200 + List", "Passed", "2026-03-08", ""],
            ["UTCID08", "Unread count", "User authenticated", "New reminds", "Returns 200 + Updated count", "Passed", "2026-03-08", ""],
            ["UTCID09", "Mark as read", "User authenticated", "Valid notif ID", "Returns 200 + Success", "Passed", "2026-03-08", ""],
            ["UTCID10", "No reminds", "System triggers", "No records due", "No notif created", "Passed", "2026-03-08", ""],
            ["UTCID11", "Pet not found", "System triggers", "Missing pet ID", "Record skipped", "Passed", "2026-03-08", ""],
        ]
        for r_idx, row in enumerate(data, 11):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value; cell.border = border; cell.alignment = center_align

    def add_update_pet_vaccination_record(wb):
        sheet = wb.create_sheet("Update Pet Vaccination Record")
        sheet['A1'] = "Function Code"; sheet['C1'] = "VAC_U01"
        sheet['A2'] = "Function Name"; sheet['C2'] = "Update Pet's Vaccination Record (PUT /vaccinations/{id})"
        sheet['A3'] = "Created By"; sheet['C3'] = "Antigravity"
        sheet['A5'] = "Lines of code"; sheet['C5'] = "116"
        sheet['A7'] = "Passed"; sheet['B7'] = "Failed"; sheet['D7'] = "Total"
        sheet['A8'] = 6; sheet['B8'] = 1; sheet['D8'] = 7
        headers = ["ID", "Condition", "Precondition", "Input Data", "Confirm (Expected Result)", "Result", "Executed Date", "Defect ID"]
        for col, text in enumerate(headers, 1):
            cell = sheet.cell(row=10, column=col)
            cell.value = text; cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
        data = [
            ["UTCID01", "Valid update (N)", "STAFF logged in", "id, valid body", "Returns 200 OK", "Passed", "2026-03-07", ""],
            ["UTCID02", "Unauthenticated (A)", "Not logged in", "id, body", "Returns 401 Unauthorized", "Passed", "2026-03-07", ""],
            ["UTCID03", "Unauthorized Role (A)", "OWNER logged in", "id, body", "Returns 403 Forbidden", "Failed", "2026-03-07", "DF012"],
            ["UTCID04", "Not Found (A)", "STAFF logged in", "Non-existent id", "Returns 404 Not Found", "Passed", "2026-03-07", ""],
            ["UTCID05", "Future Date (A)", "STAFF logged in", "Future vaccinationDate", "Returns 400 Bad Request", "Passed", "2026-03-07", ""],
            ["UTCID06", "Unsuitable Species (A)", "STAFF logged in", "Unsuitable template", "Returns 400 Bad Request", "Passed", "2026-03-07", ""],
            ["UTCID07", "Service Exception (A)", "Server error occurs", "id", "Returns 500 + Debug", "Passed", "2026-03-07", ""],
        ]
        for r_idx, row in enumerate(data, 11):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value; cell.border = border; cell.alignment = center_align

    # Add all sheets
    add_search_clinic(wb.active)
    add_delete_pet_profile(wb)
    add_create_pet_medical_record(wb)
    add_update_pet_medical_record(wb)
    add_view_pet_vaccination_record(wb)
    add_update_pet_vaccination_record(wb)
    add_create_pet_vaccination_record(wb)
    add_receive_medication_reminders(wb)
    
    wb.save("Petties_Unit_Test_Matrix.xlsx")
    print("Matrix generated: Petties_Unit_Test_Matrix.xlsx")

if __name__ == "__main__":
    create_excel_report()
